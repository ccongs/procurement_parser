"""검토 목록 장바구니 — DB CRUD + API 엔드포인트 + Excel 다운로드 테스트 — Phase 7.1.

- repository 함수(add/get/delete/clear)는 인메모리 SQLite 로 검증.
- API 엔드포인트·Excel 다운로드는 FastAPI TestClient + 임시 파일 SQLite.
실행: `pytest tests/test_export_cart.py`
"""

from __future__ import annotations

from datetime import datetime

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import repository
from app.db import Base
from app.models import AppConfig, BidNotice, ExportCartItem, PreSpec


# ---------------------------------------------------------------------------
# 공용 헬퍼
# ---------------------------------------------------------------------------
_META = datetime(2026, 6, 1, 12, 0, 0)


def _cfg() -> AppConfig:
    return AppConfig(
        id=1,
        enabled=True,
        pre_spec_enabled=True,
        auto_halted=False,
        interval_minutes=60,
        window_overlap_minutes=90,
        backfill_days=30,
        num_of_rows=20,
        max_retries=2,
        inqry_div="1",
        intrntnl_div_cd="1",
        indstryty_cds="1426,1468,1469,1470",
        updated_at=_META,
    )


def _bid(no: str, nm: str) -> BidNotice:
    return BidNotice(
        bid_ntce_no=no,
        bid_ntce_nm=nm,
        ntce_instt_nm="행정안전부",
        dminstt_nm="국세청",
        asign_bdgt_amt=50000000,
        presmpt_prce=45000000,
        bid_ntce_dt=datetime(2026, 5, 1, 9, 0),
        openg_dt=datetime(2026, 7, 1, 10, 0),
        collected_at=_META,
        updated_at=_META,
    )


def _spec(no: str, nm: str) -> PreSpec:
    return PreSpec(
        bf_spec_rgst_no=no,
        prdct_clsfc_no_nm=nm,
        order_instt_nm="조달청",
        rl_dminstt_nm="교육부",
        asign_bdgt_amt=30000000,
        rcpt_dt=datetime(2026, 5, 5, 9, 0),
        opnin_rgst_clse_dt=datetime(2026, 6, 20, 18, 0),
        sw_biz_obj_yn="Y",
        collected_at=_META,
        updated_at=_META,
    )


# ---------------------------------------------------------------------------
# 픽스처
# ---------------------------------------------------------------------------
@pytest.fixture
def session():
    """인메모리 SQLite 세션 + app_config(id=1) 시드."""
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    Local = sessionmaker(bind=engine, future=True)
    s = Local()
    s.add(_cfg())
    s.commit()
    try:
        yield s
    finally:
        s.close()


@pytest.fixture
def session_with_data(session):
    """BidNotice(BID001) + PreSpec(PS001) 시드가 추가된 세션."""
    session.add(_bid("BID001", "소프트웨어 유지보수 용역"))
    session.add(_spec("PS001", "데이터베이스 구축 사업"))
    session.commit()
    return session


@pytest.fixture
def client(tmp_path, monkeypatch):
    """임시 파일 SQLite + BidNotice/PreSpec 시드를 사용하는 TestClient."""
    from fastapi.testclient import TestClient
    from app import main

    db_path = tmp_path / "cart_test.db"
    engine = create_engine(
        f"sqlite:///{db_path}",
        future=True,
        connect_args={"check_same_thread": False},
    )
    Base.metadata.create_all(engine)
    Local = sessionmaker(bind=engine, autoflush=False, future=True)

    with Local() as s:
        s.add(_cfg())
        s.add(_bid("BID001", "소프트웨어 유지보수 용역"))
        s.add(_spec("PS001", "데이터베이스 구축 사업"))
        s.commit()

    monkeypatch.setattr(main, "SessionLocal", Local)
    return TestClient(main.app)


# ---------------------------------------------------------------------------
# repository 단위 테스트 (인메모리 SQLite)
# ---------------------------------------------------------------------------
class TestRepositoryAddAndList:
    def test_add_bid_items(self, session_with_data):
        added = repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        assert added == 1

    def test_add_pre_spec_items(self, session_with_data):
        added = repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "pre_spec", "item_id": "PS001"}],
        )
        assert added == 1

    def test_add_mixed_items(self, session_with_data):
        added = repository.add_export_cart_items(
            session_with_data,
            [
                {"item_type": "bid", "item_id": "BID001"},
                {"item_type": "pre_spec", "item_id": "PS001"},
            ],
        )
        assert added == 2

    def test_get_cart_returns_bid_meta(self, session_with_data):
        repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        items = repository.get_export_cart(session_with_data)
        assert len(items) == 1
        item = items[0]
        assert item["item_type"] == "bid"
        assert item["item_id"] == "BID001"
        assert item["title"] == "소프트웨어 유지보수 용역"
        assert item["ntce_instt_nm"] == "행정안전부"
        assert item["dminstt_nm"] == "국세청"

    def test_get_cart_returns_pre_spec_meta(self, session_with_data):
        repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "pre_spec", "item_id": "PS001"}],
        )
        items = repository.get_export_cart(session_with_data)
        assert len(items) == 1
        item = items[0]
        assert item["item_type"] == "pre_spec"
        assert item["item_id"] == "PS001"
        assert item["title"] == "데이터베이스 구축 사업"
        assert item["ntce_instt_nm"] == "조달청"
        assert item["presmpt_prce"] is None  # 사전규격은 추정가격 없음

    def test_get_cart_empty(self, session):
        items = repository.get_export_cart(session)
        assert items == []


class TestRepositoryDuplicate:
    def test_duplicate_ignored(self, session_with_data):
        repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        # 동일 아이템 재추가
        added = repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        assert added == 0

    def test_duplicate_counted_once(self, session_with_data):
        repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        items = repository.get_export_cart(session_with_data)
        assert len(items) == 1

    def test_different_type_same_id_not_duplicate(self, session):
        """item_type이 다르면 같은 item_id여도 다른 항목으로 취급."""
        session.add(_bid("SHARED001", "공통ID 입찰"))
        # PreSpec도 같은 ID로 추가(실제로는 없지만 type이 다르면 허용)
        session.add(ExportCartItem(item_type="bid", item_id="SHARED001"))
        session.add(ExportCartItem(item_type="pre_spec", item_id="SHARED001"))
        session.commit()
        items = session.query(ExportCartItem).all()
        assert len(items) == 2


class TestRepositoryDelete:
    def test_delete_by_cart_id(self, session_with_data):
        repository.add_export_cart_items(
            session_with_data,
            [{"item_type": "bid", "item_id": "BID001"}],
        )
        items = repository.get_export_cart(session_with_data)
        cart_id = items[0]["cart_id"]

        ok = repository.delete_export_cart_item(session_with_data, cart_id)
        assert ok is True

        items_after = repository.get_export_cart(session_with_data)
        assert len(items_after) == 0

    def test_delete_nonexistent_returns_false(self, session):
        ok = repository.delete_export_cart_item(session, 99999)
        assert ok is False

    def test_clear_removes_all(self, session_with_data):
        repository.add_export_cart_items(
            session_with_data,
            [
                {"item_type": "bid", "item_id": "BID001"},
                {"item_type": "pre_spec", "item_id": "PS001"},
            ],
        )
        repository.clear_export_cart(session_with_data)
        items = repository.get_export_cart(session_with_data)
        assert items == []

    def test_clear_empty_cart_no_error(self, session):
        """빈 장바구니 비우기는 에러 없이 통과."""
        repository.clear_export_cart(session)
        assert repository.get_export_cart(session) == []


# ---------------------------------------------------------------------------
# API 엔드포인트 테스트 (TestClient)
# ---------------------------------------------------------------------------
class TestApiExportCart:
    def test_get_cart_empty(self, client):
        resp = client.get("/api/export-cart")
        assert resp.status_code == 200
        data = resp.json()
        assert data["count"] == 0
        assert data["items"] == []

    def test_add_and_list(self, client):
        resp = client.post(
            "/api/export-cart",
            json={"items": [
                {"item_type": "bid", "item_id": "BID001"},
                {"item_type": "pre_spec", "item_id": "PS001"},
            ]},
        )
        assert resp.status_code == 200
        assert resp.json()["added"] == 2

        resp = client.get("/api/export-cart")
        assert resp.status_code == 200
        assert resp.json()["count"] == 2

    def test_add_duplicate_ignored(self, client):
        client.post("/api/export-cart", json={"items": [{"item_type": "bid", "item_id": "BID001"}]})
        resp = client.post("/api/export-cart", json={"items": [{"item_type": "bid", "item_id": "BID001"}]})
        assert resp.json()["added"] == 0

    def test_delete_item(self, client):
        client.post("/api/export-cart", json={"items": [{"item_type": "bid", "item_id": "BID001"}]})
        items_resp = client.get("/api/export-cart").json()["items"]
        cart_id = items_resp[0]["cart_id"]

        resp = client.delete(f"/api/export-cart/{cart_id}")
        assert resp.status_code == 200
        assert resp.json()["deleted"] is True

        assert client.get("/api/export-cart").json()["count"] == 0

    def test_delete_nonexistent_returns_false(self, client):
        resp = client.delete("/api/export-cart/99999")
        assert resp.status_code == 200
        assert resp.json()["deleted"] is False

    def test_clear_all(self, client):
        client.post("/api/export-cart", json={"items": [
            {"item_type": "bid", "item_id": "BID001"},
            {"item_type": "pre_spec", "item_id": "PS001"},
        ]})
        resp = client.delete("/api/export-cart/all")
        assert resp.status_code == 200
        assert resp.json()["cleared"] is True

        assert client.get("/api/export-cart").json()["count"] == 0

    def test_add_empty_items_list(self, client):
        resp = client.post("/api/export-cart", json={"items": []})
        assert resp.status_code == 200
        assert resp.json()["added"] == 0


class TestApiExportCartDownload:
    def test_download_empty_cart_ok(self, client):
        """장바구니가 비어 있어도 200 + spreadsheetml."""
        resp = client.get("/api/export-cart/download")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

    def test_download_with_items_ok(self, client):
        """아이템이 있을 때 200 + spreadsheetml + content-disposition."""
        client.post("/api/export-cart", json={"items": [
            {"item_type": "bid", "item_id": "BID001"},
            {"item_type": "pre_spec", "item_id": "PS001"},
        ]})
        resp = client.get("/api/export-cart/download")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "attachment" in resp.headers.get("content-disposition", "")

    def test_download_returns_valid_xlsx(self, client):
        """응답 바이트를 openpyxl로 열 수 있다."""
        from io import BytesIO
        from openpyxl import load_workbook

        client.post("/api/export-cart", json={"items": [
            {"item_type": "bid", "item_id": "BID001"},
        ]})
        resp = client.get("/api/export-cart/download")
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # 헤더 행 확인
        assert ws.cell(row=1, column=1).value == "구분"
        assert ws.cell(row=1, column=2).value == "공고번호"
        assert ws.cell(row=1, column=3).value == "제목"
        # 데이터 행 확인
        assert ws.cell(row=2, column=1).value == "입찰공고"
        assert ws.cell(row=2, column=2).value == "BID001"

    def test_download_pre_spec_row(self, client):
        """사전규격 아이템이 '사전규격'으로 표시된다."""
        from io import BytesIO
        from openpyxl import load_workbook

        client.post("/api/export-cart", json={"items": [
            {"item_type": "pre_spec", "item_id": "PS001"},
        ]})
        resp = client.get("/api/export-cart/download")
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "사전규격"
        assert ws.cell(row=2, column=2).value == "PS001"
